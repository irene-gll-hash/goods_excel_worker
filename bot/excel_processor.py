import openpyxl
from aiogram.types import Message
import asyncio
import logging
from bot.llm import get_hs_and_rates

logger = logging.getLogger(__name__)

async def _update_status(status_message: Message, text: str) -> None:
    try:
        await status_message.edit_text(text)
    except Exception as error:
        logger.warning("Не удалось обновить сообщение прогресса: %s", error)

async def process_excel(file_path: str, status_message: Message) -> tuple[str, str]:
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    data_rows = []
    for row in range(4, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if not name or str(name).strip() == "" or "пример" in str(name).lower() or "例子" in str(name):
            continue
        data_rows.append(row)
    total = len(data_rows)
    if total == 0:
        raise Exception("Не найдено ни одного товара в таблице")
    await _update_status(
        status_message,
        f"Товаров найдено: {total}\nНачинаю обработку...",
    )
    processed = 0
    successful = 0
    failed_rows = []
    heavy_count = 0
    light_count = 0

    for row in data_rows:
        name = str(ws.cell(row=row, column=1).value or "")
        purpose = str(ws.cell(row=row, column=6).value or "")
        specs = str(ws.cell(row=row, column=7).value or "")

        await _update_status(
            status_message,
            f"{processed + 1}/{total}\n<code>{name[:55]}</code>"
        )
        try:
            # === 1. Коды и ставки от DeepSeek ===
            result = await get_hs_and_rates(name, purpose, specs)
            
            cn_hs = result.get("cn_hs", "")
            ru_hs = result.get("ru_hs", "")
            duty_rate = float(result.get("duty_rate", 0)) / 100
            vat_rate = float(result.get("vat_rate", 22)) / 100

            ws.cell(row=row, column=19).value = cn_hs      # S
            ws.cell(row=row, column=20).value = ru_hs      # T
            ws.cell(row=row, column=21).value = ru_hs      # U
            ws.cell(row=row, column=22).value = duty_rate  # V
            ws.cell(row=row, column=23).value = vat_rate   # W

            # Комиссии 15% / 3%
            for col in [33, 35, 37]:
                ws.cell(row=row, column=col).value = 0.15
            for col in [34, 36, 38]:
                ws.cell(row=row, column=col).value = 0.03

            # === 2. Данные для расчётов ===
            volume = float(ws.cell(row=row, column=13).value or 0)  # M
            weight = float(ws.cell(row=row, column=14).value or 0)  # N
            qty = float(ws.cell(row=row, column=15).value or 1)     # O
            price_s = float(ws.cell(row=row, column=16).value or 0) # P
            price_m = float(ws.cell(row=row, column=17).value or 0) # Q
            price_l = float(ws.cell(row=row, column=18).value or 0) # R

            if qty <= 0:
                qty = 1

            # === 3. Определяем тип груза ПО ПЛОТНОСТИ этой строки ===
            if volume > 0:
                density = weight / volume
                is_heavy = density >= 250
            else:
                is_heavy = True  # если объём 0 — считаем тяжёлым

            if is_heavy:
                heavy_count += 1
            else:
                light_count += 1

            # === 4. Автодоставка (AC) ===
            if is_heavy:
                freight = 7 * weight / qty
            else:
                freight = 1500 * volume / qty

            ws.cell(row=row, column=29).value = round(freight, 4)

            # === 5. Таможенные пошлины (AD AE AF) ===
            def calc_duty(price, export_fee_col):
                export_fee = float(ws.cell(row=row, column=export_fee_col).value or 0.03)
                if is_heavy:
                    base = price + price * export_fee + 1.2 * weight / qty
                else:
                    base = price + price * export_fee + 300 * volume / qty
                return base * duty_rate

            duty_s = calc_duty(price_s, 34)
            duty_m = calc_duty(price_m, 36)
            duty_l = calc_duty(price_l, 38)

            ws.cell(row=row, column=30).value = round(duty_s, 6)
            ws.cell(row=row, column=31).value = round(duty_m, 6)
            ws.cell(row=row, column=32).value = round(duty_l, 6)

            # === 6. DDP (AM AN AO) ===
            def calc_ddp(price, duty, import_fee_col, export_fee_col):
                export_fee = float(ws.cell(row=row, column=export_fee_col).value or 0.03)
                import_fee = float(ws.cell(row=row, column=import_fee_col).value or 0.15)
                return (price * (1 + export_fee) + duty + freight) * (1 + import_fee) * (1 + vat_rate)

            ddp_s = calc_ddp(price_s, duty_s, 33, 34)
            ddp_m = calc_ddp(price_m, duty_m, 35, 36)
            ddp_l = calc_ddp(price_l, duty_l, 37, 38)

            ws.cell(row=row, column=39).value = round(ddp_s, 4)
            ws.cell(row=row, column=40).value = round(ddp_m, 4)
            ws.cell(row=row, column=41).value = round(ddp_l, 4)
            successful += 1

        except Exception as e:
            print(f"Ошибка в строке {row}: {e}")
            ws.cell(row=row, column=19).value = f"ERROR: {str(e)[:40]}"
            failed_rows.append(row)

        processed += 1
        await asyncio.sleep(0.35)

    result_path = file_path.replace(".xlsx", "_filled.xlsx")
    wb.save(result_path)

    info = (
        f"Обработано: {processed}\n"
        f"Успешно: {successful} | Ошибок: {len(failed_rows)}\n"
        f"Тяжёлых: {heavy_count} | Лёгких: {light_count}"
    )
    if failed_rows:
        info += f"\nСтроки с ошибками: {', '.join(map(str, failed_rows))}"
    return result_path, info
